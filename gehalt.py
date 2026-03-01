import discord
from discord.ext import commands
from discord import app_commands
from datetime import datetime, timezone
import io

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False


def dauer_formatieren(sekunden: float) -> str:
    sekunden = int(max(0, sekunden))
    h = sekunden // 3600
    m = (sekunden % 3600) // 60
    s = sekunden % 60
    teile = []
    if h > 0: teile.append(f"{h} {'Stunde' if h == 1 else 'Stunden'}")
    if m > 0: teile.append(f"{m} {'Minute' if m == 1 else 'Minuten'}")
    if s > 0 or not teile: teile.append(f"{s} {'Sekunde' if s == 1 else 'Sekunden'}")
    return ", ".join(teile)


def gehalt_berechnen(bot, member: discord.Member, nd_user: dict) -> dict:
    gc = bot.load_json(bot.SALARY_FILE)
    rc = gc.get("rollen", {})
    shift_sek   = nd_user.get("gesamt_shift_sekunden", 0)
    tickets     = nd_user.get("tickets", 0)
    shift_rate  = 0.0
    ticket_rate = 0.0
    for rolle in member.roles:
        rid = str(rolle.id)
        if rid in rc:
            shift_rate  = max(shift_rate,  float(rc[rid].get("shift_pro_stunde", 0)))
            ticket_rate = max(ticket_rate, float(rc[rid].get("ticket_bonus",     0)))
    stunden      = shift_sek / 3600
    shift_gehalt = round(stunden * shift_rate, 2)
    tick_gehalt  = round(tickets * ticket_rate, 2)
    return {
        "shift_sekunden": shift_sek,
        "shift_stunden":  round(stunden, 2),
        "tickets":        tickets,
        "shift_gehalt":   shift_gehalt,
        "ticket_gehalt":  tick_gehalt,
        "gesamt_gehalt":  round(shift_gehalt + tick_gehalt, 2),
        "shift_rate":     shift_rate,
        "ticket_rate":    ticket_rate,
    }


def gehalt_embed(guild, member: discord.Member, d: dict) -> discord.Embed:
    jetzt = datetime.now(timezone.utc)
    e = discord.Embed(color=discord.Color.from_rgb(44, 47, 51))
    if guild and guild.icon:
        e.set_author(name=guild.name, icon_url=guild.icon.url)
    e.set_thumbnail(url=member.display_avatar.url)
    e.title = "Gehaltsübersicht"
    e.description = ""
    e.add_field(name="__User__", value=f"> <:7549member:1477410605240549637> - {member.mention}", inline=False)
    e.add_field(
        name="__Übersicht Shift__",
        value=(
            f"**Zeit:** {dauer_formatieren(d['shift_sekunden'])}\n"
            f"**Satz:** {d['shift_rate']:.2f} € / Stunde\n"
            f"**Zwischensumme:** **{d['shift_gehalt']:.2f} €**"
        ),
        inline=True
    )
    e.add_field(
        name="__Übersicht Tickets__",
        value=(
            f"**Anzahl:** `{d['tickets']}`\n"
            f"**Satz:** {d['ticket_rate']:.2f} € / Ticket\n"
            f"**Zwischensumme:** **{d['ticket_gehalt']:.2f} €**"
        ),
        inline=True
    )
    e.add_field(name="__Gesamtgehalt__", value=f"<:912926arrow:1477410658806137064> {d['gesamt_gehalt']:.2f} €", inline=False)
    return e


def excel_erstellen(bot, guild: discord.Guild):
    if not EXCEL_OK:
        return None
    nd_alle = bot.load_json(bot.USERS_FILE)
    gc      = bot.load_json(bot.SALARY_FILE)
    rc      = gc.get("rollen", {})

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Gehaltsabrechnung"

    # Styles
    h_font  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    h_fill  = PatternFill(start_color="23272A", end_color="23272A", fill_type="solid")
    h_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    d_font  = Font(name="Calibri", size=10)
    d_align = Alignment(horizontal="center", vertical="center")
    alt     = PatternFill(start_color="F2F3F5", end_color="F2F3F5", fill_type="solid")
    weiss   = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    s_font  = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    s_fill  = PatternFill(start_color="7289DA", end_color="7289DA", fill_type="solid")
    wfmt    = '#,##0.00 "€"'
    hfmt    = '#,##0.00 "h"'
    duenn   = Side(style="thin", color="DDDDDD")
    rand    = Border(left=duenn, right=duenn, top=duenn, bottom=duenn)
    dick    = Border(left=duenn, right=duenn, top=duenn, bottom=Side(style="medium", color="7289DA"))

    # Titelzeile
    ws.merge_cells("A1:H1")
    jetzt_str = datetime.now(timezone.utc).strftime("%d.%m.%Y %H:%M UTC")
    c = ws["A1"]
    c.value     = f"Gehaltsabrechnung  •  {guild.name if guild else 'Server'}  •  {jetzt_str}"
    c.font      = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
    c.fill      = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Header
    headers = [
        ("Benutzer",             24),
        ("User ID",              21),
        ("Schicht (h)",          13),
        ("Schichtverdienst (€)", 22),
        ("Tickets",              11),
        ("Ticketverdienst (€)",  22),
        ("Urlaubstage",          14),
        ("Gesamtgehalt (€)",     20),
    ]
    for ci, (titel, breite) in enumerate(headers, 1):
        cell           = ws.cell(row=2, column=ci, value=titel)
        cell.font      = h_font
        cell.fill      = h_fill
        cell.alignment = h_align
        cell.border    = dick
        ws.column_dimensions[get_column_letter(ci)].width = breite
    ws.row_dimensions[2].height = 22

    # Daten
    zstart = 3
    znr    = zstart
    for uid, nd_user in nd_alle.items():
        member  = guild.get_member(int(uid)) if guild else None
        name    = nd_user.get("benutzername", f"User {uid}")
        sek     = nd_user.get("gesamt_shift_sekunden", 0)
        stunden = round(sek / 3600, 2)
        tickets = nd_user.get("tickets", 0)
        urlaub  = nd_user.get("urlaubstage", 0)

        sr = tr = 0.0
        if member:
            for rolle in member.roles:
                rid = str(rolle.id)
                if rid in rc:
                    sr = max(sr, float(rc[rid].get("shift_pro_stunde", 0)))
                    tr = max(tr, float(rc[rid].get("ticket_bonus",     0)))

        sg = round(stunden * sr, 2)
        tg = round(tickets * tr, 2)

        # Gesamtgehalt als Excel-Formel =D+F
        col_d = get_column_letter(4)
        col_f = get_column_letter(6)
        formel = f"={col_d}{znr}+{col_f}{znr}"

        fuell = alt if (znr - zstart) % 2 == 1 else weiss
        zeile = [
            (name,    d_font,                             d_align, None, fuell),
            (uid,     d_font,                             d_align, None, fuell),
            (stunden, d_font,                             d_align, hfmt, fuell),
            (sg,      d_font,                             d_align, wfmt, fuell),
            (tickets, d_font,                             d_align, None, fuell),
            (tg,      d_font,                             d_align, wfmt, fuell),
            (urlaub,  d_font,                             d_align, None, fuell),
            (formel,  Font(name="Calibri", bold=True, size=10), d_align, wfmt, fuell),
        ]
        for ci, (wert, font, align, nfmt, fill) in enumerate(zeile, 1):
            cell           = ws.cell(row=znr, column=ci, value=wert)
            cell.font      = font
            cell.alignment = align
            cell.border    = rand
            cell.fill      = fill
            if nfmt: cell.number_format = nfmt
        ws.row_dimensions[znr].height = 18
        znr += 1

    # Summenzeile
    if znr > zstart:
        ws.row_dimensions[znr].height = 6
        znr += 1
        ws.merge_cells(f"A{znr}:B{znr}")
        lc = ws.cell(row=znr, column=1, value="GESAMT")
        lc.font = s_font; lc.fill = s_fill
        lc.alignment = Alignment(horizontal="center", vertical="center")
        lc.border = rand
        for ci, nfmt in {3: hfmt, 4: wfmt, 5: None, 6: wfmt, 7: None, 8: wfmt}.items():
            buchst = get_column_letter(ci)
            cell = ws.cell(row=znr, column=ci, value=f"=SUM({buchst}{zstart}:{buchst}{znr-2})")
            cell.font = s_font; cell.fill = s_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = rand
            if nfmt: cell.number_format = nfmt
        ws.row_dimensions[znr].height = 22

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{znr}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


class ResetView(discord.ui.View):
    def __init__(self, bot):
        super().__init__(timeout=60)
        self.bot = bot

    @discord.ui.button(label="<:3518checkmark:1477410569941159959> Ja, zurücksetzen", style=discord.ButtonStyle.danger)
    async def bestaetigen(self, interaction: discord.Interaction, button: discord.ui.Button):
        if not self.bot.ist_admin(interaction.user.id):
            return await interaction.response.send_message("Fehler: `Administrator` benötigt!", ephemeral=True)
        self.bot.save_json(self.bot.SHIFTS_FILE, {})
        nd = self.bot.load_json(self.bot.USERS_FILE)
        for uid in nd:
            nd[uid]["gesamt_shift_sekunden"] = 0
            nd[uid]["shift_anzahl"]          = 0
            nd[uid]["tickets"]               = 0
        self.bot.save_json(self.bot.USERS_FILE, nd)
        e = discord.Embed(color=discord.Color.from_rgb(44, 47, 51))
        if interaction.guild and interaction.guild.icon:
            e.set_author(name=interaction.guild.name, icon_url=interaction.guild.icon.url)
        e.title       = "Gehalt zurückgesetzt!"
        e.description = "Alle Schichtzeiten und Ticket-Zähler wurden auf null gesetzt."
        await interaction.response.edit_message(embed=e, view=None)

    @discord.ui.button(label="<:3518crossmark:1477410571266818219> Abbrechen", style=discord.ButtonStyle.secondary)
    async def abbrechen(self, interaction: discord.Interaction, button: discord.ui.Button):
        e = discord.Embed(color=discord.Color.from_rgb(44, 47, 51))
        if interaction.guild and interaction.guild.icon:
            e.set_author(name=interaction.guild.name, icon_url=interaction.guild.icon.url)
        e.title = "Abgebrochen!"
        e.description = "Es wurden keine Daten verändert."
        await interaction.response.edit_message(embed=e, view=None)


class GehaltCog(commands.Cog):
    def __init__(self, bot: commands.Bot):
        self.bot          = bot
        self.gehalt_group = app_commands.Group(name="gehalt", description="Gehalts-Verwaltung")

        @self.gehalt_group.command(name="konfigurieren", description="Legt das Gehalt für eine Rolle fest (nur Admins)")
        @app_commands.describe(kategorie="Kategorie", rolle="Die Rolle", betrag="Betrag in Euro")
        @app_commands.choices(kategorie=[
            app_commands.Choice(name="Schicht (pro Stunde)", value="shift"),
            app_commands.Choice(name="Ticket (pro Ticket)",  value="ticket"),
        ])
        async def gehalt_konfigurieren(
            interaction: discord.Interaction,
            kategorie: app_commands.Choice[str],
            rolle: discord.Role,
            betrag: float
        ):
            if not bot.ist_admin(interaction.user.id):
                return await interaction.response.send_message("Fehler: `Administrator` benötigt!", ephemeral=True)
            gc = bot.load_json(bot.SALARY_FILE)
            rid = str(rolle.id)
            if rid not in gc["rollen"]:
                gc["rollen"][rid] = {"shift_pro_stunde": 0, "ticket_bonus": 0}
            if kategorie.value == "shift":
                gc["rollen"][rid]["shift_pro_stunde"] = betrag
                bez = "Schicht (pro Stunde)"
            else:
                gc["rollen"][rid]["ticket_bonus"] = betrag
                bez = "Ticket (pro Ticket)"
            bot.save_json(bot.SALARY_FILE, gc)
            e = discord.Embed(color=discord.Color.from_rgb(44, 47, 51))
            if interaction.guild and interaction.guild.icon:
                e.set_author(name=interaction.guild.name, icon_url=interaction.guild.icon.url)
            e.title = "Gehalt konfiguriert!"
            e.add_field(name="Rolle",     value=rolle.mention,        inline=True)
            e.add_field(name="Kategorie", value=bez,                  inline=True)
            e.add_field(name="Betrag",    value=f"**{betrag:.2f} €**", inline=True)
            await interaction.response.send_message(embed=e, ephemeral=True)

        @self.gehalt_group.command(name="anzeigen", description="Zeigt die Gehaltsübersicht")
        @app_commands.describe(user="(Optional) Gehalt eines anderen Users anzeigen")
        async def gehalt_anzeigen(interaction: discord.Interaction, user: discord.Member = None):
            if not bot.ist_mitarbeiter(interaction.user):
                return await interaction.response.send_message("Fehler: `Mitarbeiter` benötigt!", ephemeral=True)
            if user and user.id != interaction.user.id:
                if not bot.ist_leitungsebene(interaction.user):
                    return await interaction.response.send_message(
                        "Fehler: `Leitungsebene` benötigt!", ephemeral=True)
                ziel = user
            else:
                ziel = interaction.user
            nd  = bot.load_json(bot.USERS_FILE)
            uid = str(ziel.id)
            if uid not in nd:
                return await interaction.response.send_message(
                    f"Fehler: `Daten` für {ziel.mention} nicht gefunden!", ephemeral=True)
            d = gehalt_berechnen(bot, ziel, nd[uid])
            await interaction.response.send_message(embed=gehalt_embed(interaction.guild, ziel, d), ephemeral=True)

        @self.gehalt_group.command(name="export", description="Exportiert alle Gehaltsdaten als Excel-Datei")
        async def gehalt_export(interaction: discord.Interaction):
            if not bot.ist_leitungsebene(interaction.user):
                return await interaction.response.send_message("Fehler: `Leitungsebene` benötigt!", ephemeral=True)
            await interaction.response.defer(ephemeral=True)
            if not EXCEL_OK:
                return await interaction.followup.send("Fehler: `openpyxl` nicht installiert! Nutze: `pip install openpyxl`", ephemeral=True)
            buf = excel_erstellen(bot, interaction.guild)
            if not buf:
                return await interaction.followup.send("Fehler: `Excel-Datei` konnte nicht erstellt werden!", ephemeral=True)
            jetzt    = datetime.now(timezone.utc)
            datei    = f"gehaltsabrechnung_{jetzt.strftime('%Y%m%d_%H%M%S')}.xlsx"
            e = discord.Embed(color=discord.Color.from_rgb(44, 47, 51))
            if interaction.guild and interaction.guild.icon:
                e.set_author(name=interaction.guild.name, icon_url=interaction.guild.icon.url)
            e.title       = "Gehalt-Export"
            e.description = "Die Excel-Datei wurde erfolgreich erstellt."
            e.add_field(name="__User__", value=f"> <:7549member:1477410605240549637> - {interaction.user.mention}", inline=True)
            e.add_field(name="__Datum__", value=f"<t:{int(jetzt.timestamp())}:F>", inline=True)
            e.set_footer(text=f"Datei: {datei}")
            await interaction.followup.send(embed=e, file=discord.File(buf, filename=datei), ephemeral=True)

        @self.gehalt_group.command(name="reset", description="Setzt alle Schicht- und Ticket-Daten zurück (nur Admins)")
        async def gehalt_reset(interaction: discord.Interaction):
            if not bot.ist_admin(interaction.user.id):
                return await interaction.response.send_message("Fehler: `Administrator` benötigt!", ephemeral=True)
            e = discord.Embed(color=discord.Color.red())
            if interaction.guild and interaction.guild.icon:
                e.set_author(name=interaction.guild.name, icon_url=interaction.guild.icon.url)
            e.title       = "Gehalt zurücksetzen!"
            e.description = (
                "Dies löscht __alle__ Shift- und Ticketdaten dauerhaft.\n\n"
                "**Diese Aktion kann nicht rückgängig gemacht werden!**"
            )
            await interaction.response.send_message(embed=e, view=ResetView(bot), ephemeral=True)

    async def cog_load(self):
        self.bot.tree.add_command(self.gehalt_group)

    async def cog_unload(self):
        self.bot.tree.remove_command("gehalt")


async def setup(bot: commands.Bot):
    await bot.add_cog(GehaltCog(bot))
